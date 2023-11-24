"""Отправляет все ссылки, запрашиваемые в одно действие."""
import openpyxl 
import io
from urllib.request import urlopen

from aiogram.types import Message
from tgbot.filters.user_type import UserTypeFilter
from tgbot.models.database_instance import db


async def send_schedule(message: Message):
    """Отправляет расписание."""
    schedule_url = await db.get_schedule_url()
    if not schedule_url:
        await message.answer(text="<b>Ссылка на расписание еще не загружена менеджером!</b>")
    else:
        await message.answer(text=schedule_url)


async def send_work_schedule(message: Message):
    """Отправляет график работы."""
    work_schedule_url = await db.get_work_schedule_url()
    if not work_schedule_url:
        await message.answer(text="<b>Ссылка на график работы еще не загружена менеджером!</b>")
    else:
        await message.answer(text=work_schedule_url)


async def send_learning_schedule(message: Message):
    """Отправляет график учебы."""
    learning_schedule_url = await db.get_learning_schedule_url()
    if not learning_schedule_url:
        await message.answer(text="<b>Ссылка на график учебы еще не загружена менеджером!</b>")
    else:
        await message.answer(text=learning_schedule_url)


async def send_mailings_table(message: Message):
    """Отправляет таблицу рассылок"""
    mailings_table_url = await db.get_mailings_url()
    if not mailings_table_url:
        await message.answer(text="<b>Ссылка на таблицу рассылок еще не загружена менеджером!</b>")
    else:
        await message.answer(text=mailings_table_url)


async def send_reports(message: Message):
    """Отправляет ссылку на ведомости."""
    reports_url = await db.get_report_cards_url()
    if not reports_url:
        await message.answer(text="<b>Ссылка на ведомости еще не загружена менеджером!</b>")
    else:
        await message.answer(text=reports_url[0])


async def send_retakes_table_url(message: Message):
    """Отправляет ссылку на ведомости пересдач."""
    retakes_url = await db.get_retake_cards_url()
    if not retakes_url:
        await message.answer(text="<b>Ссылка на ведомости пересдач еще не загружена менеджером!</b>")
    else:
        await message.answer(text=retakes_url[0])


async def send_retakes_list(message: Message):
    """Отправляет список студентов, записавшихся на пересдачу."""
    retakes_url = await db.get_retake_cards_url()
    if not retakes_url:
        await message.answer(text="<b>Ссылка на ведомости пересдач еще не загружена менеджером!</b>")
    else:
        await message.answer(text="Пожалуйста, подождите. Идет обработка данных.\n" + 
                             "Это может занять несколько секунд.")
        
        file_id = retakes_url[0].split('/')[-2]
        file_url = f'https://drive.google.com/u/0/uc?id={file_id}&export=download'
        wb = openpyxl.open(filename=io.BytesIO(urlopen(file_url).read()))
        ws = wb.active
        teacher_fn = str(await db.get_user_fn(message.from_user.id)).strip("(,')")
        teacher_mn = str(await db.get_user_mn(message.from_user.id)).strip("(,')")
        teacher_ln = str(await db.get_user_ln(message.from_user.id)).strip("(,')")

        students_data = ""  
        row = 2
        while (ws.cell(row, 1).value is not None):
            if (teacher_ln == str(ws.cell(row, 10).value) and 
                teacher_fn == str(ws.cell(row, 11).value) and
                teacher_mn == str(ws.cell(row, 12).value)):

                students_data = ""
                faculty = "Факультет: " + ws.cell(row, 1).value + "\n"
                direction = "Направление: " + ws.cell(row, 2).value + "\n"
                course = "Курс: " + str(ws.cell(row, 3).value) + "\n"
                group = "Группа: " + str(ws.cell(row, 4).value) + "\n"
                student = "Студент: " + ws.cell(row, 5).value + " " +  ws.cell(row, 6).value + " " +  ws.cell(row, 7).value + "\n"
                discipline = "Предмет: " + ws.cell(row, 8).value + "\n"
                control = "Контроль: " + ws.cell(row, 9).value + "\n"
                date_of_send = "Дата отправки: " + str(ws.cell(row, 13).value)[:10] + "\n"
                students_data += faculty + direction + course + group + student + discipline + control + date_of_send + "\n"
                await message.answer(text=students_data)

            row+=1

        if (students_data == ""):
            await message.answer(text="Нет студентов, записавшихся к Вам на пересдачу.")


async def send_problem_reporting_email(message: Message):
    """Отправляет почту для жалоб."""
    problem_reporting_email = await db.get_problem_reporting_email()
    if not problem_reporting_email:
        await message.answer(text="<b>Почта поддержки еще не загружена менеджером!</b>")
    else:
        await message.answer(f"Почта поддержки: {problem_reporting_email[0]}\n\n"
                             "Пожалуйста пишите сюда, в случае обнаружения проблем в работе бота.")


def register_requested_data_sending(dp):
    dp.register_message_handler(send_schedule, ~UserTypeFilter(None), 
                                content_types=['text'], text='Расписание')
    dp.register_message_handler(send_work_schedule, UserTypeFilter("teacher"), 
                                content_types=['text'], text='График работы')
    dp.register_message_handler(send_work_schedule, UserTypeFilter("manager"), 
                                content_types=['text'], text='График работы')
    dp.register_message_handler(send_learning_schedule, UserTypeFilter("student"), 
                                content_types=['text'], text='График учебы')
    dp.register_message_handler(send_learning_schedule, UserTypeFilter("manager"), 
                                content_types=['text'], text='График учебы')
    dp.register_message_handler(send_mailings_table, UserTypeFilter("manager"), 
                                content_types=['text'], text=['Таблица рассылок'])
    dp.register_message_handler(send_mailings_table, UserTypeFilter("teacher"), 
                                content_types=['text'], text=['Таблица рассылок'])
    dp.register_message_handler(send_reports, UserTypeFilter("manager"), 
                                content_types=['text'], text=['Получить ведомости'])
    dp.register_message_handler(send_reports, UserTypeFilter("teacher"), 
                                content_types=['text'], text=['Получить ведомости'])
    dp.register_message_handler(send_retakes_table_url, UserTypeFilter("manager"), 
                                content_types=['text'], text='Получить пересдачи')
    dp.register_message_handler(send_retakes_table_url, UserTypeFilter("teacher"), 
                                content_types=['text'], text='Получить ссылку на таблицу')
    dp.register_message_handler(send_retakes_list, UserTypeFilter("teacher"), 
                                content_types=['text'], text='Получить список пересдач')
    dp.register_message_handler(send_problem_reporting_email, ~UserTypeFilter(None), 
                                content_types=['text'], text='Сообщить о проблеме')
