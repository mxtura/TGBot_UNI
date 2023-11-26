from aiogram.types import Message, CallbackQuery
from aiogram.dispatcher import FSMContext
import logging
import openpyxl
import datetime
import io
from urllib.request import urlopen
import tgbot.keyboards.reply as rkb
import tgbot.keyboards.inline as ikb
from tgbot.filters.user_type import UserTypeFilter
from tgbot.models.database_instance import db
from tgbot.misc.states import RegRetakeFSM
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload
from google.oauth2 import service_account 
import httplib2
from googleapiclient.http import MediaIoBaseUpload
from googleapiclient.discovery import build
from google.oauth2.credentials import Credentials
import io
import logging

SERVICE_ACCOUNT_FILE = 'creds.json'

# Определение областей доступа (scopes)
SCOPES = ['https://www.googleapis.com/auth/drive']

# Аутентификация и создание сервиса
credentials = service_account.Credentials.from_service_account_file(
        SERVICE_ACCOUNT_FILE, scopes=SCOPES)
service = build('drive', 'v3', credentials=credentials)


faculty_request_text = "Пожалуйста, напишите название факультета:"
direction_request_text = "Пожалуйста, напишите название направления:"
group_request_text = "Пожалуйста, напишите номер группы:"
subject_request_text = "Пожалуйста, напишите название предмета:"
control_request_text = "Пожалуйста, напишите тип контроля:"
teacher_request_text = "Пожалуйста, напишите ФИО преподавателя:"


async def getting_faculty(message: Message, state: FSMContext):
    """Просит ввести название факультета"""

    del_msg = await message.answer(text=faculty_request_text,
                                   reply_markup=rkb.faculty_input_cancel_keyboard)
    
    await state.update_data(del_msg=del_msg)
    await RegRetakeFSM.faculty.set()
    

async def getting_direction(message: Message, state: FSMContext):
    
    await state.update_data(faculty=message.text)
    
    del_msg = await message.answer(text=direction_request_text,
                                   reply_markup=rkb.direction_input_cancel_keyboard)
    await state.update_data(del_msg=del_msg)
    await RegRetakeFSM.direction.set()
    

async def getting_group(message: Message, state: FSMContext):
    
    await state.update_data(direction=message.text)  
    
    del_msg = await message.answer(text=group_request_text,
                                   reply_markup=rkb.group_input_cancel_keyboard)
    await state.update_data(del_msg=del_msg)
    await RegRetakeFSM.group.set()
    

async def getting_subject(message: Message, state: FSMContext):
    
    await state.update_data(group=message.text)  
    

    del_msg = await message.answer(text=subject_request_text,
                                   reply_markup=rkb.subject_input_cancel_keyboard)
    await state.update_data(del_msg=del_msg)
    await RegRetakeFSM.subject.set()
    

async def getting_control(message: Message, state: FSMContext):

    await state.update_data(subject=message.text)  
    
    del_msg = await message.answer(text=control_request_text,
                                   reply_markup=rkb.control_input_cancel_keyboard)

    await state.update_data(del_msg=del_msg)
    await RegRetakeFSM.control.set()    


async def getting_teacher(message: Message, state: FSMContext):

    await state.update_data(control=message.text)  
    
    
    del_msg = await message.answer(text=teacher_request_text,
                                   reply_markup=rkb.name_input_cancel_keyboard)
    await state.update_data(del_msg=del_msg)
    await RegRetakeFSM.teacher.set()
    

async def agreement(message: Message, state: FSMContext):
    await state.update_data(teacher=message.text)  

    teacher_data = await state.get_data()

    del_msg = teacher_data.get("del_msg")
    del_msg.delete()
    
    check = teacher_data["faculty"] + "\n" + teacher_data["direction"] + "\n" + teacher_data["group"] + "\n" + teacher_data["subject"] + "\n" + teacher_data["control"] + "\n" + teacher_data["teacher"]
    
    msgs_to_del = [await message.answer(text=f"Правильно? \n<b>{check}</b>",
                                            reply_markup=rkb.confirmation_keyboard)]
    await state.update_data(msgs_to_del=msgs_to_del)

    await RegRetakeFSM.agreement.set()


async def write_retakes_list(message: Message, state: FSMContext):
    """Записывает данные о пересдачах в excel-файл на google drive"""

    teacher_data = await state.get_data()
    
    student_fn = str(await db.get_user_fn(message.from_user.id)).strip("(,')")
    student_mn = str(await db.get_user_mn(message.from_user.id)).strip("(,')")
    student_ln = str(await db.get_user_ln(message.from_user.id)).strip("(,')")

    retakes_url = await db.get_retake_cards_url()
    file_id = retakes_url[0].split('/')[-2]
    file_url = f'https://drive.google.com/u/0/uc?id={file_id}&export=download'
    fn = io.BytesIO(urlopen(file_url).read())
    wb = openpyxl.load_workbook(fn)
    
    ws = wb.active

    msgs_to_del = teacher_data.get("msgs_to_del")
    for msg in msgs_to_del:
        await msg.delete()

    # Находим первую пустую строку
    row = 1
    while ws.cell(row, 1).value is not None:
        row += 1

    # Записываем данные в найденную строку
    current_date = datetime.datetime.now().date().strftime("%d.%m.%Y")

    data = [teacher_data["faculty"], teacher_data["direction"], teacher_data["group"][0], await db.get_user_group_name(message.from_user.id), student_ln, student_fn, student_mn, teacher_data["subject"], teacher_data["control"], teacher_data["teacher"].split()[0], teacher_data["teacher"].split()[1], teacher_data["teacher"].split()[2], current_date]
    for col, value in enumerate(data, start=1):
        ws.cell(row, col).value = value


    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)

    # Загрузка обновленного файла обратно на Google Drive
    media = MediaIoBaseUpload(excel_buffer, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    service.files().update(fileId=file_id, media_body=media).execute()

    await message.answer("Данные успешно записаны")



def register_reg_retake(dp):
    dp.register_message_handler(getting_faculty, UserTypeFilter("student"), content_types=['text'],
                                text=['Записаться на пересдачу'])
    dp.register_message_handler(getting_direction, state=RegRetakeFSM.faculty)
    dp.register_message_handler(getting_group, state=RegRetakeFSM.direction)
    dp.register_message_handler(getting_subject, state=RegRetakeFSM.group)
    dp.register_message_handler(getting_control, state=RegRetakeFSM.subject)
    dp.register_message_handler(getting_teacher, state=RegRetakeFSM.control)
    dp.register_message_handler(agreement, state=RegRetakeFSM.teacher)
    dp.register_message_handler(write_retakes_list, state=RegRetakeFSM.agreement)
    
